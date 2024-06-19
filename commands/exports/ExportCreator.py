import coc
import disnake
import openpyxl.worksheet.worksheet
import io
import calendar
from datetime import datetime, timedelta
from disnake.ext import commands
from typing import TYPE_CHECKING
from classes.bot import CustomClient
from classes.player.stats import StatsPlayer, LegendDay
from typing import Dict, List
from coc import utils
from openpyxl import load_workbook, Workbook
from pytz import utc


class ExportCreator(commands.Cog):
    def __init__(self, bot: CustomClient):
        self.bot = bot

    async def write_data(
        self,
        worksheet: openpyxl.worksheet.worksheet.Worksheet,
        column_names: List[str],
        data: List[List],
    ):
        col = 1
        data.insert(0, column_names)
        for row, d in enumerate(data, start=1):
            for item in d:
                worksheet.cell(row=row, column=col).value = item
                col += 1
            col = 1
        return worksheet

    async def export_manager(
        self, player_tags: List[str], season: str = None, template: str = None
    ):
        # get list of custom players (which have lots of db info), use the cache since not time sensitive
        players: List[StatsPlayer] = await self.bot.get_players(
            tags=player_tags, custom=True, use_cache=True
        )
        output = io.BytesIO()
        # if the "template" is just the name of a default type (raw data), just export the 1 sheet
        if template in self.DEFAULT_EXPORT_TYPES:
            workbook = Workbook()
            workbook.remove_sheet(workbook["Sheet"])
            if template == "Legend Stats":
                await self.create_legend_export(
                    players=players,
                    workbook=workbook,
                    season=season,
                    sheet_name="legend_stats",
                )
            elif template == "War Hits":
                await self.create_warhit_export(
                    players=players,
                    workbook=workbook,
                    season=season,
                    sheet_name="war_hits",
                )
            elif template == "Season Trophies":
                await self.create_season_trophies_export(
                    players=players,
                    workbook=workbook,
                    season=season,
                    sheet_name="season_trophies",
                )
            elif template == "Troops":
                await self.create_troops_export(
                    players=players, workbook=workbook, sheet_name="season_troops"
                )
            elif template == "Player Achievements":
                await self.create_achievements_export(
                    players=players, workbook=workbook, sheet_name="player_achievements"
                )
            elif template == "Player Activity":
                await self.create_player_activity_export(
                    players=players,
                    workbook=workbook,
                    season=season,
                    sheet_name="player_activity",
                )
            elif template == "Player Stats":
                await self.create_player_stats_export(
                    players=players,
                    workbook=workbook,
                    season=season,
                    sheet_name="player_stats",
                )
            elif template == "Advanced Stats":
                await self.create_advanced_player_stats_export(
                    players=players,
                    workbook=workbook,
                    season=season,
                    sheet_name="advanced_stats",
                )
        else:
            # if it is not, then it is a template
            # 1. load the template
            # 2. look for sheet names that match export types so they can be removed & replaced with an updated version
            # 3. look if they have a number to find what season that is being exported, else it is just the current
            workbook = load_workbook(template)
            for sheet_name in workbook.sheetnames:
                print(sheet_name)
                season_for_sheet = season
                # this code assumes that all export type names are 2 parts seperated by underscore. if this is different, then other logic can apply
                # i.e. could split & check if the last item is an integer
                # also as a sidenote, a lot of functions of mine, assume that if "season" is None, then it defaults to the current season
                # whether this is in helper functions, core code, or even button mechanics for users
                if len(sheet_name.split("_")) == 3:
                    season_spot = int(sheet_name.split("_")[-1])
                    # generate this number of seasons
                    # since we generate the *exact amount* the one we need will always be the last one
                    season_for_sheet = self.bot.gen_season_date(season_spot)[-1]
                    # however this returns it as Month Year & we need YYYY-MM
                    # not convenient, but we have written the code once before (exports.py - season convertor)
                    # could skip this all by writting a season generator that actually gives the right thing, if u feel inclined xD
                    # or we could switch all generators to give back datetimes which would allow us to create whatever we want with them...hindsight is 20/20 lol
                    month = list(calendar.month_name).index(
                        season_for_sheet.split(" ")[0]
                    )
                    year = int(season_for_sheet.split(" ")[1])
                    if month == 1:
                        month = 13
                        year -= 1
                    end_date = coc.utils.get_season_end(month=int(month - 1), year=year)
                    month = end_date.month
                    if month <= 9:
                        month = f"0{month}"
                    season_for_sheet = f"{end_date.year}-{month}"
                if "legend_stats" in sheet_name:
                    workbook.remove(workbook.get_sheet_by_name(sheet_name))
                    await self.create_legend_export(
                        players=players,
                        workbook=workbook,
                        season=season_for_sheet,
                        sheet_name=sheet_name,
                    )
                elif "war_hits" in sheet_name:
                    workbook.remove(workbook.get_sheet_by_name(sheet_name))
                    await self.create_warhit_export(
                        players=players,
                        workbook=workbook,
                        season=season_for_sheet,
                        sheet_name=sheet_name,
                    )
                elif "season_trophies" in sheet_name:
                    workbook.remove(workbook.get_sheet_by_name(sheet_name))
                    await self.create_season_trophies_export(
                        players=players,
                        workbook=workbook,
                        season=season_for_sheet,
                        sheet_name=sheet_name,
                    )
                elif "season_troops" in sheet_name:
                    workbook.remove(workbook.get_sheet_by_name(sheet_name))
                    await self.create_troops_export(
                        players=players, workbook=workbook, sheet_name=sheet_name
                    )
                elif "player_achievements" in sheet_name:
                    workbook.remove(workbook.get_sheet_by_name(sheet_name))
                    await self.create_achievements_export(
                        players=players, workbook=workbook, sheet_name=sheet_name
                    )
                elif "player_activity" in sheet_name:
                    workbook.remove(workbook.get_sheet_by_name(sheet_name))
                    await self.create_player_activity_export(
                        players=players,
                        workbook=workbook,
                        season=season_for_sheet,
                        sheet_name=sheet_name,
                    )
                elif "player_stats" in sheet_name:
                    workbook.remove(workbook.get_sheet_by_name(sheet_name))
                    await self.create_player_stats_export(
                        players=players, workbook=workbook, sheet_name=sheet_name
                    )
                elif "advanced_stats" in sheet_name:
                    workbook.remove(workbook.get_sheet_by_name(sheet_name))
                    await self.create_advanced_player_stats_export(
                        players=players,
                        workbook=workbook,
                        season=season_for_sheet,
                        sheet_name=sheet_name,
                    )
                # more if statements to find other export types
        workbook.save(output)
        xlsx_data = output
        xlsx_data.seek(0)
        return xlsx_data

    async def create_advanced_player_stats_export(
        self,
        players: List[StatsPlayer],
        workbook: openpyxl.Workbook,
        sheet_name: str,
        season: str = None,
    ):
        advanced_player_Stats_page = workbook.create_sheet(sheet_name)
        year = season[:4]
        month = season[-2:]
        SEASON_START = utils.get_season_start(month=int(month) - 1, year=int(year))
        SEASON_END = utils.get_season_end(month=int(month) - 1, year=int(year))
        weeks = []
        SEASON_START = SEASON_START - timedelta(3)
        for i in range(0, 7):
            week = SEASON_START + timedelta(i * 7)
            if week > SEASON_END:
                break
            weeks.append(week)
        players_data = await self.bot.get_players(
            tags=[player.tag for player in players], custom=True
        )
        if season is None:
            season = self.bot.gen_season_date()
        data = []
        for player in players_data:
            donos = player.donos(season)
            capital_raided = 0
            capital_donated = 0
            for date in weeks:
                capital = player.clan_capital_stats(week=str(date.date()))
                capital_raided += sum(capital.raided)
                capital_donated += sum(capital.donated)

            if player.last_online is None:
                lastOnline = "-"
            else:
                lastOnline = datetime.fromtimestamp(
                    player.last_online, tz=utc
                ).strftime("%Y-%m-%d-%H:%M:%S")
            data.append(
                [
                    player.name,
                    player.tag,
                    lastOnline,
                    len(player.season_last_online(season)),
                    player.attack_wins,
                    donos.received,
                    donos.donated,
                    player.gold_looted(season),
                    player.elixir_looted(season),
                    player.dark_elixir_looted(season),
                    capital_raided,
                    capital_donated,
                ]
            )

        columns = [
            "Player Name",
            "Player Tag",
            "Last Online",
            "Activity Count",
            "Attack Wins",
            "Received",
            "Donated",
            "Gold Looted",
            "Elixir Looted",
            "Dark Elixir Looted",
            "Capital Raided",
            "Capital Donated",
        ]
        await self.write_data(
            worksheet=advanced_player_Stats_page, column_names=columns, data=data
        )

    async def create_player_stats_export(
        self, players: List[StatsPlayer], workbook: openpyxl.Workbook, sheet_name: str
    ):
        player_stats_page = workbook.create_sheet(sheet_name)
        players_data = await self.bot.get_players(
            tags=[player.tag for player in players], custom=False
        )
        data = []
        for player in players_data:
            info = [
                "name",
                "tag",
                "town_hall",
                "town_hall_weapon",
                "exp_level",
                "trophies",
                "best_trophies",
                "labels",
                "league",
                "donations",
                "received",
                "attack_wins",
                "defense_wins",
                "clan",
                "role",
                "war_opted_in",
                "war_stars",
                "builder_hall",
                "best_versus_trophies",
                "clan_capital_contributions",
                "versus_trophies",
            ]
            line = []
            for i in info:
                value = getattr(player, i, None)
                value = value.name if i == "league" else value
                value = "-" if value is None or value == "" else value
                if i == "labels":
                    labels = [label.name for label in value] + [
                        "-" for x in range(3 - len(value))
                    ]
                    line.extend(labels)
                else:
                    line.append(str(value))
            data.append(line)
        columns = [
            "Player Name",
            "Player Tag",
            "Town Hall",
            "Town Hall Weapon",
            "Exp Level",
            "Trophies",
            "Best Trophies",
            "Label 1",
            "Label 2",
            "Label 3",
            "League",
            "Donations",
            "Received",
            "Attack Wins",
            "Defense Wins",
            "Clan",
            "Role",
            "War Opted In",
            "War Stars",
            "Builder Hall",
            "Best Versus Trophies",
            "Clan Capital Contributions",
            "Versus Trophies",
        ]

        await self.write_data(
            worksheet=player_stats_page, column_names=columns, data=data
        )

    async def create_troops_export(
        self, players: List[StatsPlayer], workbook: openpyxl.Workbook, sheet_name: str
    ):
        troops_page = workbook.create_sheet(sheet_name)
        players_data = await self.bot.get_players(
            tags=[player.tag for player in players], custom=False
        )
        # Note home_troop_order already includes siege machines
        home_troops = coc.enums.HOME_TROOP_ORDER
        super_troops = coc.enums.SUPER_TROOP_ORDER
        spells = coc.enums.SPELL_ORDER
        pets = coc.enums.PETS_ORDER
        home_heroes = coc.enums.HOME_BASE_HERO_ORDER
        builder_troops = coc.enums.BUILDER_TROOPS_ORDER
        builder_heroes = coc.enums.BUILDER_BASE_HERO_ORDER

        columns = (
            home_troops
            + super_troops
            + spells
            + pets
            + home_heroes
            + builder_troops
            + builder_heroes
        )
        data = []
        for player in players_data:
            _troops = {troop.name: troop.level for troop in player.troops}
            _spells = {spell.name: spell.level for spell in player.spells}
            _super_troops = {
                super_troop.name: super_troop.level
                for super_troop in player.super_troops
            }
            _pets = {pet.name: pet.level for pet in player.pets}
            _heros = {hero.name: hero.level for hero in player.heroes}
            all_troops_data = {**_troops, **_spells, **_super_troops, **_pets, **_heros}
            line = [player.name, player.tag] + [
                all_troops_data.get(column, "-") for column in columns
            ]
            data.append(line)
        columns = ["Player Name", "Player Tag"] + columns
        await self.write_data(worksheet=troops_page, column_names=columns, data=data)

    async def create_player_activity_export(
        self,
        players: List[StatsPlayer],
        workbook: openpyxl.Workbook,
        sheet_name: str,
        season: str = None,
    ):
        activity_page = workbook.create_sheet(sheet_name)
        year = season[:4]
        month = season[-2:]
        SEASON_START = utils.get_season_start(
            month=int(month) - 1, year=int(year)
        ).timestamp()
        SEASON_END = utils.get_season_end(
            month=int(month) - 1, year=int(year)
        ).timestamp()
        pipeline = [
            {
                "$match": {
                    "$and": [
                        {"tag": {"$in": [player.tag for player in players]}},
                        {"time": {"$gte": SEASON_START}},
                        {"time": {"$lte": SEASON_END}},
                    ]
                }
            },
            {"$sort": {"type": 1}},
            {
                "$group": {
                    "_id": "$tag",
                    "changes": {
                        "$push": {
                            "type": "$type",
                            "clan": "$clan",
                            "time": "$time",
                            "value": "$value",
                            "p_value": "$p_value",
                        }
                    },
                }
            },
            {
                "$lookup": {
                    "from": "player_stats",
                    "localField": "_id",
                    "foreignField": "tag",
                    "as": "name",
                }
            },
            {"$set": {"name": "$name.name"}},
        ]
        results: List[dict] = await self.bot.player_history.aggregate(pipeline).to_list(
            length=None
        )
        data = []
        leagues = ["builderBaseLeague", "league"]
        for result in results:
            for change in result["changes"]:
                p_value = change.get("p_value", "-")
                value = change["value"]
                if change["type"] in leagues:
                    value = change["value"]["name"]
                    p_value = change["p_value"]["name"] if p_value != "-" else p_value
                time = datetime.fromtimestamp(change["time"], tz=utc).strftime(
                    "%Y-%m-%d-%H:%M:%S"
                )
                data.append(
                    [
                        result["name"][0],
                        result["_id"],
                        change["type"],
                        p_value,
                        value,
                        time,
                        change["clan"],
                    ]
                )
        columns = [
            "Player Name",
            "Player Tag",
            "Type",
            "Previous Value",
            "Value",
            "Time",
            "Clan Tag",
        ]
        await self.write_data(worksheet=activity_page, column_names=columns, data=data)

    async def create_achievements_export(
        self, players: List[StatsPlayer], workbook: openpyxl.workbook, sheet_name: str
    ):
        achievement_page = workbook.create_sheet(sheet_name)
        players_data = await self.bot.get_players(
            tags=[player.tag for player in players], custom=False
        )
        data = []
        achievement_order = coc.enums.ACHIEVEMENT_ORDER
        for player in players_data:
            achievements = []
            entry = [player.name, player.tag]
            for achievement in player.achievements:
                achievements.append(achievement.name)
                entry.append(achievement.value)
            if "Get those Goblins!" not in achievements:
                entry.insert(33, "-")
            data.append(entry)
        columns = ["Player Name", "Player Tag"] + achievement_order
        columns.remove("Get those other Goblins!")
        # columns.remove("Get even more Goblins!")
        await self.write_data(
            worksheet=achievement_page, column_names=columns, data=data
        )
        return workbook

    async def create_season_trophies_export(
        self,
        players: List[StatsPlayer],
        workbook: openpyxl.Workbook,
        sheet_name: str,
        season: str = None,
    ):
        season_trophies_page = workbook.create_sheet(sheet_name)
        trophies_data = await self.bot.history_db.find(
            {
                "$and": [
                    {"tag": {"$in": [player.tag for player in players]}},
                    {"season": season},
                ]
            }
        ).to_list(length=None)
        data = [
            [
                entry["name"],
                entry["tag"],
                entry["expLevel"],
                entry["trophies"],
                entry["attackWins"],
                entry["defenseWins"],
                entry["rank"],
                entry["clan"]["name"],
                entry["clan"]["tag"],
                entry["season"],
            ]
            for entry in trophies_data
        ]

        columns = [
            "Player Name",
            "Player Tag",
            "Exp Level",
            "Trophies",
            "Attack Wins",
            "Defense Wins",
            "Rank",
            "Clan Name",
            "Clan Tag",
            "Season",
        ]

        await self.write_data(
            worksheet=season_trophies_page, column_names=columns, data=data
        )
        return workbook

    async def create_warhit_export(
        self,
        players: List[StatsPlayer],
        workbook: openpyxl.Workbook,
        sheet_name: str,
        season: str = None,
    ):
        warhit_stat_page = workbook.create_sheet(sheet_name)
        year = season[:4]
        month = season[-2:]
        SEASON_START = utils.get_season_start(
            month=int(month) - 1, year=int(year)
        ).timestamp()
        SEASON_END = utils.get_season_end(
            month=int(month) - 1, year=int(year)
        ).timestamp()
        attacks = await self.bot.warhits.find(
            {
                "$and": [
                    {"tag": {"$in": [player.tag for player in players]}},
                    {"_time": {"$gte": SEASON_START}},
                    {"_time": {"$lte": SEASON_END}},
                ]
            }
        ).to_list(length=None)

        defends = await self.bot.warhits.find(
            {
                "$and": [
                    {"defender_tag": {"$in": [player.tag for player in players]}},
                    {"_time": {"$gte": SEASON_START}},
                    {"_time": {"$lte": SEASON_END}},
                ]
            }
        ).to_list(length=None)
        data = []
        for attack in attacks:
            data.append(
                [
                    "Attack",
                    attack["name"],
                    attack["tag"],
                    attack["townhall"],
                    datetime.fromtimestamp(attack["_time"], tz=utc).strftime(
                        "%Y-%m-%d-%H:%M:%S"
                    ),
                    attack["destruction"],
                    attack["stars"],
                    attack["fresh"],
                    datetime.fromtimestamp(attack["war_start"], tz=utc).strftime(
                        "%Y-%m-%d-%H:%M:%S"
                    ),
                    attack["defender_tag"],
                    attack["defender_name"],
                    attack["defender_townhall"],
                    attack["war_type"],
                    attack["war_status"],
                    attack["attack_order"],
                    attack["map_position"],
                    attack.get("war_size", 0),
                    attack.get("clan", "No Clan"),
                ]
            )

        for defense in defends:
            data.append(
                [
                    "Defense",
                    defense["name"],
                    defense["tag"],
                    defense["townhall"],
                    datetime.fromtimestamp(defense["_time"], tz=utc).strftime(
                        "%Y-%m-%d-%H:%M:%S"
                    ),
                    defense["destruction"],
                    defense["stars"],
                    defense["fresh"],
                    datetime.fromtimestamp(defense["war_start"], tz=utc).strftime(
                        "%Y-%m-%d-%H:%M:%S"
                    ),
                    defense["defender_tag"],
                    defense["defender_name"],
                    defense["defender_townhall"],
                    defense["war_type"],
                    defense["war_status"],
                    defense["attack_order"],
                    defense["map_position"],
                    defense.get("war_size", 0),
                    defense.get("clan", "No Clan"),
                ]
            )

        columns = [
            "Hit Type",
            "Player Name",
            "Player Tag",
            "Townhall",
            "Time",
            "Destruction",
            "Stars",
            "Fresh",
            "War Start",
            "Defender Tag",
            "Defender Name",
            "Defender Townhall",
            "War Type",
            "War Status",
            "Attack Order",
            "Map Position",
            "War Size",
            "Clan",
        ]

        await self.write_data(
            worksheet=warhit_stat_page, column_names=columns, data=data
        )
        return workbook

    async def create_legend_export(
        self,
        players: List[StatsPlayer],
        workbook: openpyxl.Workbook,
        sheet_name: str,
        season: str = None,
    ):
        legend_stats_page = workbook.create_sheet(sheet_name)
        start = utils.get_season_start().replace(tzinfo=utc).date()
        now = datetime.now(tz=utc).date()
        current_season_progress = now - start
        current_season_progress = current_season_progress.days
        if season != self.bot.gen_season_date():
            current_season_progress = 100
        data = []
        for player in players:
            season_stats: Dict[str, LegendDay] = player.season_of_legends(season=season)
            day_spot = 0
            for day, legend_day in season_stats.items():
                day_spot += 1
                data.append(
                    [
                        player.name,
                        player.tag,
                        player.clan_name(),
                        player.clan_tag(),
                        day,
                        legend_day.attack_sum,
                        legend_day.defense_sum,
                        legend_day.net_gain,
                        legend_day.num_attacks.integer,
                        legend_day.num_defenses.integer,
                    ]
                )
                if day_spot == current_season_progress:
                    break
        columns = [
            "Player Name",
            "Player Tag",
            "Clan Name",
            "Clan Tag",
            "Day",
            "Attack Sum",
            "Defense Sum",
            "Net Gain",
            "Num Attacks",
            "Num Defenses",
        ]

        await self.write_data(
            worksheet=legend_stats_page, column_names=columns, data=data
        )
        return workbook

    """#THESE ARE JUST PROTOTYPES, MAY HAVE SOME GOOD STUFF, MAY NOT.
    async def create_last_season_trophies_export(self, ctx, clan):
        workbook = xlsxwriter.Workbook(f'{clan.tag}_last_season_end.xlsx', {'in_memory' : True})
        worksheet = workbook.add_worksheet("Legend_Trophies")

        background_color = "white"
        letter_color = "black"
        bold = workbook.add_format(
            {"font_color": letter_color, 'bold': True, "bg_color": background_color, "align": "center"})
        center = workbook.add_format({"font_color": letter_color, "align": "center", "bg_color": background_color})
        perc_hr = workbook.add_format({"font_color": letter_color, 'num_format': '0.0%', "bg_color": background_color})
        blue_letters = workbook.add_format({"font_color": "blue", "bg_color": background_color})
        white_back = workbook.add_format({"font_color": letter_color, "bg_color": background_color})
        gold_letters = workbook.add_format({"align": "center", "font_color": "orange", "bg_color": background_color})
        row = 0
        col = 0

        types = ["Rank", "TH", "Name", "Trophies"]

        length_list = [len(x) for x in types]
        for i, width in enumerate(length_list):
            worksheet.set_column(i, i, width)

        all_players_info = [types]
        clan_members: list[coc.Player] = await self.bot.get_players(tags=[member.tag for member in clan.members])

        sort_list = []
        for player in clan_members:
            try:
                id = player.legend_statistics.previous_season.id
                start = utils.get_season_start().replace(tzinfo=tiz).date()
                mon = start.month
                if mon <= 9:
                    mon = f"0{mon}"
                season = f"{start.year}-{mon}"
                if str(id) != season:
                    continue
                legend_trophies = player.legend_statistics.previous_season.trophies
                sort_list.append([player.name, player, legend_trophies])
            except:
                continue

        sort_list = sorted(sort_list, key=lambda l: (-l[-1], l[0]), reverse=False)
        clan_members = [member[1] for member in sort_list]

        longest_name = 0
        for rank, player in enumerate(clan_members, 1):
            player_info = [rank, player.town_hall, player.name, player.legend_statistics.previous_season.trophies]
            longest_name = max(longest_name, len(player.name))
            all_players_info.append(player_info)

        worksheet.set_column(2, 2, longest_name + 1)
        for data in all_players_info:
            for item in data:
                if row != 0:
                    if col == 0:
                        worksheet.write(row, col, item, gold_letters)
                    elif col == 1:
                        worksheet.write(row, col, item, blue_letters)
                    elif col == 2:
                        worksheet.write(row, col, item, white_back)
                    elif col == 3:
                        worksheet.write(row, col, item, center)
                else:
                    worksheet.write(row, col, item, bold)
                col += 1
            row += 1
            col = 0

        workbook.close()
        file = disnake.File(f'ExportInfo/{clan.tag}_last_season_end.xlsx', filename=f"{clan.name}_last_season_end.xlsx")
        excel2img.export_img(f'ExportInfo/{clan.tag}_last_season_end.xlsx',
                             f"ExportInfo/{clan.tag}_last_season_end.png", "",
                             f"Legend_Trophies!A1:D{min(len(clan_members) + 1, 50)}")
        image = disnake.File(f'ExportInfo/{clan.tag}_last_season_end.png', filename=f"text.png")
        await ctx.edit_original_message(files=[image, file])

    async def create_member_export(self, ctx, clan):
        num = random.randint(1, 10)
        workbook = xlsxwriter.Workbook(f'member_export{num}.xlsx')
        worksheet = workbook.add_worksheet()
        bold = workbook.add_format({'bold': True})

        row = 0
        col = 0

        types = ["Name", "TH"]
        for hero in coc.HERO_ORDER:
            types.append(hero)
        for troop in coc.HOME_TROOP_ORDER:
            types.append(troop)

        length_list = [len(x) for x in types]
        for i, width in enumerate(length_list):
            worksheet.set_column(i, i, width)

        all_players_info = [types]
        clan_members: list[coc.Player] = await self.bot.get_players(tags=[member.tag for member in clan.members])
        longest_name = 0
        for player in clan_members:
            player_info = [player.name, player.town_hall]
            longest_name = max(longest_name, len(player.name))
            for hero_name in coc.HERO_ORDER:
                hero = player.get_hero(name=hero_name)
                if hero is None:
                    hero_level = 0
                else:
                    hero_level = hero.level
                player_info.append(hero_level)

            for troop_name in coc.HOME_TROOP_ORDER:
                troop = player.get_troop(name=troop_name)
                if troop is None:
                    troop_level = 0
                else:
                    troop_level = troop.level
                player_info.append(troop_level)

            all_players_info.append(player_info)

        worksheet.set_column(0, 0, longest_name + 1)
        for data in all_players_info:
            for item in data:
                if col != 0 and row != 0:
                    worksheet.write(row, col, item)
                else:
                    worksheet.write(row, col, item, bold)
                col += 1
            row += 1
            col = 0

        workbook.close()
        file = disnake.File(f'member_export{num}.xlsx', filename=f"{clan.name}_member_stats.xlsx")
        await ctx.edit_original_message(file=file)

    async def create_war_export(self, clan):
        workbook = xlsxwriter.Workbook(f'ExportInfo/war_stats{clan.tag}.xlsx')

        clan_members: list[StatsPlayer] = await self.bot.get_players(tags=[member.tag for member in clan.members],
                                                                        custom=True)
        worksheet = workbook.add_worksheet("Hitrates")

        await self.create_war_hitrate(clan_members=clan_members, workbook=workbook, worksheet=worksheet)
        await self.create_war_defrate(clan_members=clan_members, workbook=workbook, worksheet=worksheet)
        await self.create_war_starlb(clan_members=clan_members, workbook=workbook, worksheet=worksheet)

        workbook.close()
        file = disnake.File(f'ExportInfo/war_stats{clan.tag}.xlsx', filename=f"{clan.name}_war_stats.xlsx")
        excel2img.export_img(f'ExportInfo/war_stats{clan.tag}.xlsx', f"ExportInfo/war_stats{clan.tag}.png", "",
                             f"Hitrates!A1:T{min(len(clan_members) + 1, 26)}")
        image = disnake.File(f'ExportInfo/war_stats{clan.tag}.png', filename=f"test.png")
        return (image, file)

    async def create_war_hitrate(self, clan_members, workbook, worksheet):
        background_color = "white"
        letter_color = "black"
        bold = workbook.add_format(
            {"font_color": letter_color, 'bold': True, "bg_color": background_color, "align": "center"})
        center = workbook.add_format({"font_color": letter_color, "align": "center", "bg_color": background_color})
        perc_hr = workbook.add_format({"font_color": letter_color, 'num_format': '0.0%', "bg_color": background_color})
        blue_letters = workbook.add_format({"font_color": "blue", "bg_color": background_color})
        white_back = workbook.add_format({"font_color": letter_color, "bg_color": background_color})
        gold_letters = workbook.add_format({"align": "center", "font_color": "orange", "bg_color": background_color})
        row = 0
        col = 0

        types = ["Rank", "TH", "Name", "  Total  ", "  3★  ", "Total Stars"]

        length_list = [len(x) for x in types]
        for i, width in enumerate(length_list):
            worksheet.set_column(i, i, width)

        all_players_info = [types]

        sort_list = []
        player_hr = {}
        for player in clan_members:
            hitrate = await player.hit_rate()
            hitrate = hitrate[0]

            sort_list.append(
                [player.name, player, hitrate.average_triples])
            player_hr[player.tag] = hitrate

        hitrate_sort_list = sorted(sort_list, key=lambda l: (-l[-1], l[0]), reverse=False)
        hitrate_clan_members = [member[1] for member in hitrate_sort_list]

        longest_name = 0
        for rank, player in enumerate(hitrate_clan_members, 1):
            hitrate = player_hr[player.tag]
            player_info = [rank, player.town_hall, player.name, f"{hitrate.total_triples}/{hitrate.num_attacks}",
                           hitrate.average_triples, hitrate.total_stars]
            longest_name = max(longest_name, len(player.name))
            all_players_info.append(player_info)

        worksheet.set_column(2, 2, longest_name + 1)
        for data in all_players_info:
            for item in data:
                if row != 0:
                    if col == 0:
                        worksheet.write(row, col, item, gold_letters)
                    elif col == 1:
                        worksheet.write(row, col, item, blue_letters)
                    elif col == 2:
                        worksheet.write(row, col, item, white_back)
                    elif col == 3:
                        worksheet.write(row, col, item, center)
                    elif col == 4:
                        worksheet.write(row, col, item, perc_hr)
                    elif col == 5:
                        worksheet.write(row, col, item, center)
                else:
                    worksheet.write(row, col, item, bold)
                col += 1
            row += 1
            col = 0

    async def create_war_defrate(self, clan_members, workbook, worksheet):
        background_color = "white"
        letter_color = "black"
        bold = workbook.add_format(
            {"font_color": letter_color, 'bold': True, "bg_color": background_color, "align": "center"})
        center = workbook.add_format({"font_color": letter_color, "align": "center", "bg_color": background_color})
        perc_hr = workbook.add_format({"font_color": letter_color, 'num_format': '0.0%', "bg_color": background_color})
        blue_letters = workbook.add_format({"font_color": "blue", "bg_color": background_color})
        white_back = workbook.add_format({"font_color": letter_color, "bg_color": background_color})
        gold_letters = workbook.add_format({"align": "center", "font_color": "orange", "bg_color": background_color})
        row = 0
        col = 7
        og_col = 7

        types = ["Rank", "TH", "Name", "  Total  ", "  DR  ", "Total Stars"]

        length_list = [len(x) for x in types]
        for i, width in enumerate(length_list):
            worksheet.set_column(i + og_col, i + og_col, width)

        all_players_info = [types]

        sort_list = []
        player_dr = {}
        for player in clan_members:
            defense_rate = await player.defense_rate()
            defense_rate = defense_rate[0]
            if defense_rate.num_attacks == 0:
                continue
            sort_list.append(
                [player.name, player, defense_rate.average_triples,
                 round(100 - (defense_rate.average_triples * 100), 1)])
            player_dr[player.tag] = defense_rate

        def_rate_sort_list = sorted(sort_list, key=lambda l: (-l[-2], l[0]), reverse=False)
        def_rate_clan_members = [member[1] for member in def_rate_sort_list]

        longest_name = 0
        for rank, player in enumerate(def_rate_clan_members, 1):
            def_rate = player_dr[player.tag]
            player_info = [rank, player.town_hall, player.name, f"{def_rate.total_triples}/{def_rate.num_attacks}",
                           def_rate.average_triples, def_rate.total_stars]
            longest_name = max(longest_name, len(player.name))
            all_players_info.append(player_info)

        worksheet.set_column(og_col + 2, og_col + 2, longest_name + 1)
        for data in all_players_info:
            for item in data:
                if row != 0:
                    if col == og_col:
                        worksheet.write(row, col, item, gold_letters)
                    elif col == og_col + 1:
                        worksheet.write(row, col, item, blue_letters)
                    elif col == og_col + 2:
                        worksheet.write(row, col, item, white_back)
                    elif col == og_col + 3:
                        worksheet.write(row, col, item, center)
                    elif col == og_col + 4:
                        worksheet.write(row, col, item, perc_hr)
                    elif col == og_col + 5:
                        worksheet.write(row, col, item, center)
                else:
                    worksheet.write(row, col, item, bold)
                col += 1
            row += 1
            col = og_col

    async def create_war_starlb(self, clan_members, workbook, worksheet):
        background_color = "white"
        letter_color = "black"
        bold = workbook.add_format(
            {"font_color": letter_color, 'bold': True, "bg_color": background_color, "align": "center"})
        center = workbook.add_format({"font_color": letter_color, "align": "center", "bg_color": background_color})
        perc_hr = workbook.add_format({"font_color": letter_color, 'num_format': '0.0%', "bg_color": background_color})
        blue_letters = workbook.add_format({"font_color": "blue", "bg_color": background_color})
        white_back = workbook.add_format({"font_color": letter_color, "bg_color": background_color})
        gold_letters = workbook.add_format({"align": "center", "font_color": "orange", "bg_color": background_color})
        row = 0
        col = 14
        og_col = 14

        types = ["Rank", "TH", "Name", "Stars", "Attacks", "Triples"]

        length_list = [len(x) for x in types]
        for i, width in enumerate(length_list):
            worksheet.set_column(i + og_col, i + og_col, width)

        all_players_info = [types]

        sort_list = []
        player_hr = {}
        for player in clan_members:
            hit_rate = await player.hit_rate()
            hit_rate = hit_rate[0]

            sort_list.append(
                [player.name, player, hit_rate.total_stars])
            player_hr[player.tag] = hit_rate

        star_sort_list = sorted(sort_list, key=lambda l: (-l[-1], l[0]), reverse=False)
        star_clan_members = [member[1] for member in star_sort_list]

        longest_name = 0
        for rank, player in enumerate(star_clan_members, 1):
            hit_rate = player_hr[player.tag]
            player_info = [rank, player.town_hall, player.name, hit_rate.total_stars, hit_rate.num_attacks,
                           hit_rate.total_triples]
            longest_name = max(longest_name, len(player.name))
            all_players_info.append(player_info)

        worksheet.set_column(og_col + 2, og_col + 2, longest_name + 1)
        for data in all_players_info:
            for item in data:
                if row != 0:
                    if col == og_col:
                        worksheet.write(row, col, item, gold_letters)
                    elif col == og_col + 1:
                        worksheet.write(row, col, item, blue_letters)
                    elif col == og_col + 2:
                        worksheet.write(row, col, item, white_back)
                    elif col == og_col + 3:
                        worksheet.write(row, col, item, center)
                    elif col == og_col + 4:
                        worksheet.write(row, col, item, center)
                    elif col == og_col + 5:
                        worksheet.write(row, col, item, center)
                else:
                    worksheet.write(row, col, item, bold)
                col += 1
            row += 1
            col = og_col"""
